from django.conf import settings
from django.http import HttpResponse
import cStringIO as StringIO
import openpyxl
from openpyxl.workbook import Workbook
from openpyxl.writer.excel import save_virtual_workbook
from openpyxl.cell import get_column_letter
import re
from ecwsp.sis.helper_functions import strip_unicode_to_ascii

class XlReport:
    """ Wrapper for openpyxl
    Using xlsx because xls has limitations and ods has no good python
    library
    """
    def __init__(self, file_name="Report"):
        """ file_name does not need an extention """
        if file_name.endswith('.xls'):
            file_name = file_name[:-4]
        elif file_name.endswith('xlsx'):
            file_name = file_name[:-5]
        file_name = file_name.replace(' ', '_') # Some browsers don't deal well with spaces in downloads
        self.workbook = Workbook()
        self.workbook.remove_sheet(self.workbook.get_active_sheet())
        self.file_name = file_name
        # Sniff the openpyxl version
        try:
            from openpyxl import __major__ as openpyxl_major_version
            self.old_openpyxl = openpyxl_major_version < 2
        except ImportError:
            self.old_openpyxl = False
    
    def add_sheet(self, data, title=None, header_row=None, heading=None, auto_width=False, max_auto_width=50):
        """ Add a sheet with data to workbook
        title: sheet name
        header_row: List - Column header (bold with bottom border)
        heading: Sheet heading (very top of sheet)
        auto_width: will ESTIMATE the width of each column by counting
        max chars in each column. It will not work with a formula.
        max_auto_width: is the max number of characters a column to be
        """
        sheet = self.workbook.create_sheet()
        if title:
            # Maximum 31 characters allowed in sheet title
            clean_title = title[:31]
            # From openpyxl's _set_title()
            bad_title_char_re = re.compile(r'[\\*?:/\[\]]')
            # Replace bad characters with underscores
            clean_title = bad_title_char_re.sub('_', clean_title)
            sheet.title = unicode(clean_title)
        if heading:
            sheet.append([unicode(heading)])
        if header_row:
            header_row = map(unicode, header_row)
            sheet.append(header_row)
            row = sheet.get_highest_row()
            for i, header_cell in enumerate(header_row):
                if self.old_openpyxl:
                    cell = sheet.cell(row=row-1, column=i)
                    cell.style.font.bold = True
                    cell.style.borders.bottom.border_style = openpyxl.style.Border.BORDER_THIN
                else:
                    cell = sheet.cell(row=row, column=i+1)
                    cell.style = cell.style.copy(
                        font=cell.style.font.copy(bold=True),
                        border=openpyxl.styles.Border(
                            bottom=openpyxl.styles.Side(
                                border_style=openpyxl.styles.borders.BORDER_THIN
                            )
                        )
                    )
        for row in data:
            row = map(unicode, row)
            sheet.append(row)
        if auto_width:
            column_widths = []
            for row in data:
                row = map(unicode, row)
                for i, cell in enumerate(row):
                    if len(column_widths) > i:
                        if len(cell) > column_widths[i]:
                            column_widths[i] = len(cell)
                    else:
                        column_widths += [len(cell)]
            
            for i, column_width in enumerate(column_widths):
                if column_width > 3:
                    if column_width < max_auto_width:
                        # * 0.9 estimates a typical variable width font
                        sheet.column_dimensions[get_column_letter(i+1)].width = column_width * 0.9
                    else:
                        sheet.column_dimensions[get_column_letter(i+1)].width = max_auto_width
    
    def save(self, filename):
        self.workbook.save(settings.MEDIA_ROOT + filename)
    
    def as_download(self):
        """ Returns a django HttpResponse with the xlsx file """
        myfile = StringIO.StringIO()
        myfile.write(save_virtual_workbook(self.workbook))
        response = HttpResponse(
            myfile.getvalue(),
            content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')

        safe_filename = re.sub(
            '[^A-Za-z0-9]+',
            '_',
            strip_unicode_to_ascii(self.file_name)
        )
        response['Content-Disposition'] = 'attachment; filename=%s.xlsx' % safe_filename
        response['Content-Length'] = myfile.tell()
        return response
