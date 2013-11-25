#   Copyright 2011 Burke Software and Consulting LLC
#   Author David M Burke <david@burkesoftware.com>
#   
# This program is free software; you can redistribute it and/or modify
# it under the terms of the GNU General Public License as published by
# the Free Software Foundation; either version 3 of the License, or
# (at your option) any later version.
#
# This program is distributed in the hope that it will be useful,
# but WITHOUT ANY WARRANTY; without even the implied warranty of
# MERCHANTABILITY or FITNESS FOR A PARTICULAR PURPOSE. See the
# GNU General Public License for more details.
#
# You should have received a copy of the GNU General Public License
# along with this program. If not, see <http://www.gnu.org/licenses/>.

from django.utils.html import strip_tags
from django.core.exceptions import ObjectDoesNotExist
from django.db.models import F, Sum

from ecwsp.sis.report import *
from ecwsp.sis.helper_functions import Struct
from ecwsp.sis.template_report import TemplateReport
from ecwsp.sis.models import Cohort
from ecwsp.administration.models import Template
from ecwsp.omr.models import AnswerInstance
from ecwsp.benchmarks.models import Benchmark

import xlwt
from ecwsp.sis.xl_report import XlReport
from openpyxl.cell import get_column_letter

class ReportManager(object):
    def download_results(self, test):
        """ Create basic xls report for OMR. Includes summary and details """
        
        # from download_teacher_results()
        total_test_takers = test.active_testinstance_set.filter(answerinstance__points_earned__gt=0).distinct().count()

        # Summary sheet
        data = [[test.name]]
        data.append(["Points Possible:", test.points_possible])
        data.append(["Results collected: %s" % (test.students_test_results,)])
        data.append(['Test average: %s' % (test.points_average,)])
        data.append([''])
        data.append(['Student', 'Points Earned', 'Percentage'])
        first_student_row = i = 7
        for ti in test.active_testinstance_set.annotate(earned=Sum('answerinstance__points_earned')):
            data.append([ti.student, ti.earned, "=B%s / $B$2" % i])
            i += 1
        # Make it easier to compare this against download_teacher_results()
        data.append([''])
        i += 1
        data.append(["Students scoring at or above 70%",
            # number of points possible is in cell B2
            '=COUNTIF(B{0}:B{1},">="&0.7*B2)'.format(first_student_row, i - 2),
            # don't put the decimal inside the string to avoid localization problems
            '=COUNTIF(C{0}:C{1},">="&0.7)/COUNT(C6:C{1})'.format(first_student_row, i - 2)])
        i += 1
        
        report = XlReport(file_name="OMR Report")
        report.add_sheet(data, title="Summary", auto_width=True)
        # Detail sheets
        data_points = []
        data_answers = []
        data_abc = []
        row_points = ["Student"]
        row_answers = ["Student"]
        row_abc = ["Student"]
        for i,question in enumerate(test.question_set.all()):
            row_points.append("%s %s" % (question.get_order_start_one, strip_tags(question.question).strip()))
            row_answers.append("%s %s" % (question.get_order_start_one, strip_tags(question.question).strip()))
            row_abc.append("Question {0}".format(i+1))
        data_points.append(row_points)
        data_answers.append(row_answers)
        data_abc.append(row_abc)
        
        for test_instance in test.active_testinstance_set.all():
            row_points = []
            row_answers = []
            row_abc = []
            row_points.append(test_instance.student)
            row_answers.append(test_instance.student)
            row_abc.append(test_instance.student)
            for question in test.question_set.all():
                try:
                    answer = test_instance.answerinstance_set.get(question=question)
                    row_points.append(answer.points_earned)
                    row_answers.append(strip_tags(answer.answer).strip())
                    i = None
                    if question.type == "True/False":
                        row_abc += [answer.answer]
                    else:
                        for i,x in enumerate(question.answer_set.all()):
                            if x == answer.answer:
                                break
                        row_abc += [chr(65+i)]
                except ObjectDoesNotExist:
                    row_points += ['']
                    row_answers += ['']
                    row_abc += ['']
            data_points.append(row_points)
            data_answers.append(row_answers)
            data_abc.append(row_abc)
        
        report.add_sheet(data_points, title="Detail Points", auto_width=True)
        report.add_sheet(data_answers, title="Detail Answers", auto_width=True)
        report.add_sheet(data_abc, title="Answer Sheet", auto_width=True)
        
        # Benchmark sheet
        data = []
        row = ['Benchmark']
        row2 = ['Points Possible']
        benchmarks = Benchmark.objects.filter(question__test=test).distinct()
        for benchmark in benchmarks:
            row.append(benchmark)
            row.append('%')
            row2.append(test.question_set.filter(benchmarks=benchmark).aggregate(Sum('point_value'))['point_value__sum'])
            row2.append('')
        data.append(row)
        data.append(row2)
        first_student_row = i = 3 # 3 for third row on spreadsheet
        for test_instance in test.active_testinstance_set.all():
            row = [test_instance.student]
            a = 2 # the letter b or column b in spreadsheet
            for benchmark in Benchmark.objects.filter(question__test=test).distinct():
                row.append(test_instance.answerinstance_set.filter(
                    question__benchmarks=benchmark).aggregate(
                    Sum('points_earned'))['points_earned__sum'])
                row.append('={0}{1}/{0}$2'.format(get_column_letter(a), str(i)))
                a += 2 # skip ahead 2 columns
            i += 1
            data.append(row)
        # Make it easier to compare this against download_teacher_results()
        data.append([''])
        i += 1
        row = ['Students scoring at or above 70%']
        col = 2
        while col < a:
            row.append('=COUNTIF({0}{1}:{0}{2},">="&0.7*{0}{3})'.format(
                get_column_letter(col), first_student_row, i - 2, first_student_row - 1))
            col += 1
            row.append('=COUNTIF({0}{1}:{0}{2},">="&0.7)/COUNT({0}{1}:{0}{2})'.format(
                get_column_letter(col), first_student_row, i - 2))
            col += 1
        data.append(row)
        i += 1
        report.add_sheet(data, title="Benchmark", auto_width=True)
        
        data = [['Benchmark', 'Name', 'Earned', 'Possible', 'Percent']]
        i = 2
        for benchmark in benchmarks:
            row = []
            row += [benchmark.number, benchmark.name]
            answer_data = AnswerInstance.objects.filter(
                test_instance__in=test.active_testinstance_set.all(),
                question__benchmarks=benchmark).aggregate(
                    Sum('points_earned'),
                    Sum('points_possible'))
            row += [answer_data['points_earned__sum']]
            # this causes a discrepancy with download_teacher_results() if
            # a student leaves a question blank.
            #row += [answer_data['points_possible__sum']]
            # instead, get the points possible for all questions having this benchmark and
            # multiply by the number of test takers
            benchmark_points_possible = test.question_set.filter(benchmarks=benchmark).aggregate(
                Sum('point_value'))['point_value__sum'] * total_test_takers
            row.append(benchmark_points_possible)
            row += ['=C{0}/D{0}'.format(str(i))]
            data += [row]
            i += 1
        report.add_sheet(data, title="Benchmarks for class", auto_width=True)
        
        return report.as_download()
        
    def download_teacher_results(self, test, format, template, cohorts=None):
        """ Make appy based report showing results for a whole class """
        if not cohorts:
            cohorts = Cohort.objects.all()
        
        # Stupid fucking hack
        subquery = test.testinstance_set.filter(student__cohort__in=cohorts).distinct()

        report = TemplateReport()
        report.file_format = format
        test_instances = test.active_testinstance_set.filter(answerinstance__points_earned__gt=0).filter(pk__in=subquery).annotate(Sum('answerinstance__points_earned'))
        test.benchmarks = Benchmark.objects.filter(question__test=test).distinct()
        
        points_possible = test.points_possible
        points_to_earn = 0.70 * test.points_possible
        number_gte_70 = test_instances.filter(pk__in=subquery).filter(answerinstance__points_earned__sum__gte=points_to_earn).count()
        total_test_takers = test_instances.filter(pk__in=subquery).filter(answerinstance__points_earned__gt=0).distinct().count()
        try:
            test.percent_gte_70 = float(number_gte_70) / total_test_takers
        except ZeroDivisionError:
            test.percent_gte_70 = 0.0
        test.report_average = test.get_average(cohorts=cohorts)

        for benchmark in test.benchmarks:
            # TODO: eliminate this subquery? is the idea to eliminate any questions that all students left unanswered?
            qb_subquery = test.question_set.filter(answerinstance__test_instance__student__cohort__in=cohorts).distinct()
            question_benchmarks = test.question_set.filter(pk__in=qb_subquery).filter(benchmarks=benchmark).distinct()
            benchmark.points_possible = question_benchmarks.aggregate(Sum('point_value'))['point_value__sum']
            benchmark.total_points_possible = benchmark.points_possible * test_instances.count()
            # Really think this should work...but nope.
            #benchmark.total_points_earned = question_benchmarks.aggregate(Sum('answerinstance__points_earned'))['answerinstance__points_earned__sum']
            earned_sum = 0
            for question_benchmark in question_benchmarks:
                for answer in question_benchmark.active_answerinstance_set.filter(test_instance__student__cohort__in=cohorts).distinct():
                    earned_sum += answer.points_earned
            benchmark.total_points_earned = earned_sum

            benchmark.average = float(benchmark.total_points_earned) / benchmark.total_points_possible 
           
            # Percent of students scoring at or above 70%
            test_instances_gte_70 = 0
            for test_instance in test_instances:
                 answers = test_instance.answerinstance_set.filter(question__benchmarks=benchmark)
                 answers_points = answers.aggregate(Sum('points_earned'), Sum('points_possible'))
                 instance_points_earned = answers_points['points_earned__sum']
                 instance_points_possible = answers_points['points_possible__sum']
                 if instance_points_earned and instance_points_possible:
                     instance_average = float(instance_points_earned) / instance_points_possible
                     if instance_average >= 0.70:
                         test_instances_gte_70 += 1
            benchmark.gte_70 = float(test_instances_gte_70) / test_instances.count()

            benchmark.assessed_on = ""
            for question_benchmark in question_benchmarks:
                benchmark.assessed_on += "{}, ".format(question_benchmark.get_order_start_one)
            benchmark.assessed_on = benchmark.assessed_on[:-2]
        
        test.questions = test.question_set.all()
        for question in test.questions:
            question.benchmark_text = ''
            for benchmark in question.benchmarks.all():
                question.benchmark_text += '{}, '.format(benchmark.number)
            question.benchmark_text = question.benchmark_text[:-2]
            # grab all the AnswerInstances that we care about for this question
            answerinstances = question.answerinstance_set.filter(test_instance__student__cohort__in=cohorts).distinct()
            # nasty! http://stackoverflow.com/questions/4093910/django-aggregates-sums-in-postgresql-dont-use-distinct-is-this-a-bug/4917507#4917507 
            answerinstances = question.active_answerinstance_set.filter(pk__in=answerinstances)
            # calculate the COUNT of correct student responses for this question
            question.num_correct = answerinstances.filter(points_earned__gte=F('points_possible')).count()
            # calculate the COUNT of all student responses for this question
            question.num_total = answerinstances.count()
            # http://www.merriam-webster.com/dictionary/percent: "cent" means 100, but I'll stick with the existing convention 
            question.percent_correct = float(question.num_correct) / question.num_total
            # calculate the sum of all points earned and the sum of all points possible for this question
            earned_possible = answerinstances.aggregate(Sum('points_earned'), Sum('points_possible'))
            question.points_earned = earned_possible['points_earned__sum'] 
            question.points_possible = earned_possible['points_possible__sum'] 
            question.percent_points_earned = float(question.points_earned) / question.points_possible
            
        report.data['test'] = test
        report.data['tests'] = test_instances
        
        report.filename = 'Teacher Results for ' + unicode(test)
        return report.pod_save(template)  
        
    def download_student_results(self, test, format, template):
        """ Make appy based report showing results for each student """
        report = TemplateReport()
        report.file_format = format
        test_instances = test.active_testinstance_set.all()
        benchmarks = Benchmark.objects.filter(question__test=test).distinct()
        
        for benchmark in benchmarks:
            benchmark.points_possible = test.question_set.filter(benchmarks=benchmark).aggregate(Sum('point_value'))['point_value__sum']
        
        for test_instance in test_instances:
            benchmark_instances = []
            for benchmark in benchmarks:
                benchmark_instance = Struct()
                benchmark_instance.benchmark = benchmark
                benchmark_instance.points_possible = benchmark.points_possible
                benchmark_instance.answers = test_instance.answerinstance_set.filter(question__benchmarks=benchmark)
                benchmark_instance.points_earned = benchmark_instance.answers.aggregate(Sum('points_earned'))['points_earned__sum']
                benchmark_instance.questions = ''
                for answer in benchmark_instance.answers.all():
                    benchmark_instance.questions += '{}, '.format(answer.question.get_order_start_one)
                benchmark_instance.questions = benchmark_instance.questions[:-2]
                benchmark_instances.append(benchmark_instance)
            test_instance.benchmarks = benchmark_instances
        
            test_instance.incorrects = test_instance.answerinstance_set.filter(points_earned__lt=F('points_possible'))
            for incorrect in test_instance.incorrects:
                incorrect.benchmarks = ''
                for benchmark in incorrect.question.benchmarks.all():
                    incorrect.benchmarks += '{}, '.format(benchmark.number)
                incorrect.benchmarks = incorrect.benchmarks[:-2]
                
                try:
                    incorrect.right_answer = incorrect.question.answer_set.order_by('point_value').reverse()[0]
                except:
                    incorrect.right_answer = "No correct answer"
            
        report.data['test'] = test
        report.data['tests'] = test_instances
        
        report.filename = 'Student Results for ' + unicode(test)
        return report.pod_save(template)  

report = ReportManager()
